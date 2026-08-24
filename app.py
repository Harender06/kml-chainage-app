import io
import math
import xml.etree.ElementTree as ET
from geopy.distance import geodesic
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import simplekml
import streamlit as st

st.set_page_config(
    page_title="Highway Best Fit Alignment & Chainage Tool", layout="wide"
)

st.title("🛣️ Highway Best Fit Alignment & Chainage Engine")
st.markdown(
    "Upload raw KML to generate **Best Fit Alignment**, automatic **Chainage Markers**, **Voice Alert Tour**, and an **Excel Engineering Report**."
)

# ---------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------


def extract_coords(kml_bytes):
    root = ET.fromstring(kml_bytes)
    coords = []
    for elem in root.iter():
        if elem.tag.endswith("coordinates"):
            raw_text = elem.text.strip()
            points = raw_text.split()
            for p in points:
                parts = p.split(",")
                if len(parts) >= 2:
                    lon, lat = float(parts[0]), float(parts[1])
                    coords.append((lat, lon))
            if coords:
                break
    return coords


def fit_best_tangents_rdp(pts, max_deviation_m):
    """Segment points into best fit straight tangents with max allowed offset deviation in meters"""
    if len(pts) < 1:
        return pts

    lat_deg_to_m = 111000.0
    lon_deg_to_m = 111000.0 * math.cos(math.radians(pts[0][0]))

    dmax = 0.0
    index = 0
    end = len(pts) - 1

    p1 = np.array(pts[0])
    p2 = np.array(pts[end])

    p1_m = np.array([p1[0] * lat_deg_to_m, p1[1] * lon_deg_to_m])
    p2_m = np.array([p2[0] * lat_deg_to_m, p2[1] * lon_deg_to_m])

    for i in range(1, end):
        p3 = np.array(pts[i])
        p3_m = np.array([p3[0] * lat_deg_to_m, p3[1] * lon_deg_to_m])

        if np.all(p1_m == p2_m):
            d = np.linalg.norm(p3_m - p1_m)
        else:
            v1 = p2_m - p1_m
            v2 = p1_m - p3_m
            cross_val = abs(v1[0] * v2[1] - v1[1] * v2[0])
            d = cross_val / np.linalg.norm(v1)

        if d > dmax:
            index = i
            dmax = d

    if dmax > max_deviation_m:
        rec_res1 = fit_best_tangents_rdp(pts[: index + 1], max_deviation_m)
        rec_res2 = fit_best_tangents_rdp(pts[index:], max_deviation_m)
        return rec_res1[:-1] + rec_res2
    else:
        return [pts[0], pts[end]]


def calculate_bearing(p1, p2):
    lat1, lon1 = math.radians(p1[0]), math.radians(p1[1])
    lat2, lon2 = math.radians(p2[0]), math.radians(p2[1])
    dlon = lon2 - lon1
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(
        lat2
    ) * math.cos(dlon)
    initial_bearing = math.atan2(x, y)
    return (math.degrees(initial_bearing) + 360) % 360


def generate_voice_tour_kml(smoothed_coords, start_chainage, interval_km=1.0):
    kml_tour = simplekml.Kml(name="Alignment Voice Navigation")

    line = kml_tour.newlinestring(
        name="Route Line", coords=[(c[1], c[0]) for c in smoothed_coords]
    )
    line.style.linestyle.width = 4
    line.style.linestyle.color = simplekml.Color.cyan

    tour = kml_tour.newgxtour(name="▶️ Start Voice Navigation Tour")
    playlist = tour.newgxplaylist()

    accumulated_dist = 0.0
    curr_ch = float(start_chainage)
    target_km = (
        math.ceil(curr_ch / 1000.0) * 1000.0
        if curr_ch % 1000 != 0
        else curr_ch + 1000.0
    )

    for i in range(len(smoothed_coords) - 1):
        p1 = smoothed_coords[i]
        p2 = smoothed_coords[i + 1]
        seg_m = geodesic(p1, p2).meters

        while (curr_ch + accumulated_dist + seg_m) >= target_km:
            overshoot = target_km - (curr_ch + accumulated_dist)
            frac = overshoot / seg_m if seg_m > 0 else 0

            t_lat = p1[0] + frac * (p2[0] - p1[0])
            t_lon = p1[1] + frac * (p2[1] - p1[1])

            km_val = int(target_km // 1000)
            alert_text = f"Chainage {km_val} Kilometer"

            pnt = kml_tour.newpoint(
                name=f"🔊 {km_val}+000 KM", coords=[(t_lon, t_lat)]
            )
            pnt.description = f"Voice Alert: {alert_text}"
            pnt.style.iconstyle.scale = 1.2
            pnt.style.iconstyle.color = simplekml.Color.blue

            flyto = playlist.newgxflyto()
            flyto.camera.latitude = t_lat
            flyto.camera.longitude = t_lon
            flyto.camera.altitude = 150
            flyto.camera.tilt = 45

            wait = playlist.newgxwait()
            wait.duration = 2.0

            target_km += interval_km * 1000.0

        accumulated_dist += seg_m

    return kml_tour.kml()


def generate_excel_alignment_report(
    curve_data, total_len, design_speed, terrain_type
):
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    section_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws_sum["A1"] = "HIGHWAY BEST FIT ALIGNMENT DESIGN REPORT"
    ws_sum["A1"].font = title_font

    ws_sum["A3"] = "1. Project & Alignment Summary"
    ws_sum.merge_cells("A3:D3")
    ws_sum["A3"].font = bold_font
    ws_sum["A3"].fill = section_fill

    summary_rows = [
        ("Design Speed", f"{design_speed} km/h"),
        ("Terrain Classification", terrain_type),
        ("Total Alignment Length", f"{total_len/1000:.3f} km"),
        ("Total Best Fit Curves/PIs", len(curve_data)),
        ("Design Code Standard", "IRC:73 / IRC:37 / MORTH"),
    ]

    for r_idx, (k, v) in enumerate(summary_rows, start=4):
        ws_sum.cell(row=r_idx, column=1, value=k).font = bold_font
        ws_sum.cell(row=r_idx, column=2, value=v).font = normal_font

    ws_curves = wb.create_sheet(title="Horizontal Curve Schedule")
    ws_curves.views.sheetView[0].showGridLines = True

    headers = [
        "PI No.",
        "Latitude (°)",
        "Longitude (°)",
        "Deflection Angle Δ (°)",
        "Direction",
        "Fitted Radius R (m)",
        "Tangent Length T (m)",
        "Curve Length L (m)",
        "PC Chainage (m)",
        "PI Chainage (m)",
        "PT Chainage (m)",
    ]

    ws_curves.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws_curves.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, cd in enumerate(curve_data, start=2):
        direction = "Right" if cd["Deflection Angle (deg)"] > 0 else "Left"
        row_vals = [
            cd["PI Index"],
            cd["Latitude"],
            cd["Longitude"],
            abs(cd["Deflection Angle (deg)"]),
            direction,
            cd["Design Radius (m)"],
            cd["Tangent Length T (m)"],
            cd["Curve Length L (m)"],
            cd["PC Chainage (m)"],
            cd["PI Chainage (m)"],
            cd["PT Chainage (m)"],
        ]
        ws_curves.append(row_vals)

        for c_idx in range(1, len(row_vals) + 1):
            cell = ws_curves.cell(row=row_idx, column=c_idx)
            cell.font = normal_font
            cell.border = thin_border
            if c_idx in [2, 3, 4, 6, 7, 8, 9, 10, 11]:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    for ws in [ws_sum, ws_curves]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------
# MAIN INTERFACE & PROCESSING
# ---------------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload Existing Road KML File", type=["kml"], key="main_kml"
)

st.markdown("#### ⚙️ Alignment & Chainage Settings")

col1, col2, col3, col4 = st.columns(4)
with col1:
    max_allowed_offset = st.slider(
        "Max Allowed Offset (m)",
        min_value=1.0,
        max_value=15.0,
        value=3.0,
        step=0.5,
    )
with col2:
    min_curve_radius = st.number_input(
        "Min Radius R (m)",
        value=100.0,
        step=10.0,
    )
with col3:
    start_chainage = st.number_input(
        "Start Chainage (m)", value=0, step=100
    )
with col4:
    design_speed = st.number_input(
        "Design Speed (km/h)",
        value=60,
        step=10,
    )

col_ch1, col_ch2, col_ch3 = st.columns(3)
with col_ch1:
    major_interval = st.number_input("Major Interval (m)", value=100, step=10)
with col_ch2:
    minor_interval = st.number_input("Minor Interval (m)", value=20, step=5)
with col_ch3:
    reverse_direction = st.checkbox(
        "🔄 Reverse Direction (Start chainage from opposite end)"
    )

output_name = st.text_input(
    "Output KML Name (Optional)",
    value="Alignment_With_Chainages",
)

if uploaded_file is not None:
    if st.button(
        "🚀 Generate Alignment + Chainages + Voice Tour", type="primary"
    ):
        raw_coords = extract_coords(uploaded_file.read())

        if not raw_coords or len(raw_coords) < 3:
            st.error("Insufficient points in KML!")
        else:
            if reverse_direction:
                raw_coords = raw_coords[::-1]

            pi_coords = fit_best_tangents_rdp(raw_coords, max_allowed_offset)

            curve_data = []
            smoothed_coords = [(pi_coords[0][0], pi_coords[0][1])]

            kml_fit = simplekml.Kml()

            orig_line = kml_fit.newlinestring(
                name="Original Ground Track",
                coords=[(c[1], c[0]) for c in raw_coords],
            )
            orig_line.style.linestyle.width = 2
            orig_line.style.linestyle.color = simplekml.Color.yellow

            running_chainage = 0.0

            for i in range(1, len(pi_coords) - 1):
                p_prev = pi_coords[i - 1]
                p_curr = pi_coords[i]
                p_next = pi_coords[i + 1]

                b1 = calculate_bearing(p_prev, p_curr)
                b2 = calculate_bearing(p_curr, p_next)

                deflection = b2 - b1
                if deflection > 180:
                    deflection -= 360
                elif deflection < -180:
                    deflection += 360

                abs_def = abs(deflection)
                delta_rad = math.radians(abs_def)

                dist_prev = geodesic(p_prev, p_curr).meters
                dist_next = geodesic(p_curr, p_next).meters

                desired_t = (
                    min_curve_radius * math.tan(delta_rad / 2)
                    if delta_rad > 0
                    else 0
                )
                max_t = min(dist_prev / 2, dist_next / 2)

                actual_t = min(desired_t, max_t)
                fitted_radius = (
                    (actual_t / math.tan(delta_rad / 2))
                    if delta_rad > 0
                    else min_curve_radius
                )
                arc_length = fitted_radius * delta_rad

                frac_pc = 1.0 - (actual_t / dist_prev if dist_prev > 0 else 0)
                pc_lat = p_prev[0] + frac_pc * (p_curr[0] - p_prev[0])
                pc_lon = p_prev[1] + frac_pc * (p_curr[1] - p_prev[1])

                frac_pt = actual_t / dist_next if dist_next > 0 else 0
                pt_lat = p_curr[0] + frac_pt * (p_next[0] - p_curr[0])
                pt_lon = p_curr[1] + frac_pt * (p_next[1] - p_curr[1])

                dist_to_pi = geodesic(p_prev, p_curr).meters
                pi_chainage = running_chainage + dist_to_pi
                pc_chainage = pi_chainage - actual_t
                pt_chainage = pc_chainage + arc_length

                curve_data.append(
                    {
                        "PI Index": f"PI-{i}",
                        "Latitude": round(p_curr[0], 6),
                        "Longitude": round(p_curr[1], 6),
                        "Deflection Angle (deg)": round(deflection, 2),
                        "Design Radius (m)": round(fitted_radius, 1),
                        "Tangent Length T (m)": round(actual_t, 2),
                        "Curve Length L (m)": round(arc_length, 2),
                        "PC Chainage (m)": round(pc_chainage, 2),
                        "PI Chainage (m)": round(pi_chainage, 2),
                        "PT Chainage (m)": round(pt_chainage, 2),
                    }
                )

                arc_points = 12
                for step in range(arc_points + 1):
                    f = step / arc_points
                    lat_interp = (1 - f) ** 2 * pc_lat + 2 * (
                        1 - f
                    ) * f * p_curr[0] + f**2 * pt_lat
                    lon_interp = (1 - f) ** 2 * pc_lon + 2 * (
                        1 - f
                    ) * f * p_curr[1] + f**2 * pt_lon
                    smoothed_coords.append((lat_interp, lon_interp))

                running_chainage = pt_chainage

            smoothed_coords.append((pi_coords[-1][0], pi_coords[-1][1]))

            fit_line = kml_fit.newlinestring(
                name="Best Fit Road Alignment",
                coords=[(c[1], c[0]) for c in smoothed_coords],
            )
            fit_line.style.linestyle.width = 4
            fit_line.style.linestyle.color = simplekml.Color.cyan

            accumulated_dist = 0.0
            curr_ch = float(start_chainage)

            km = int(curr_ch // 1000)
            m = int(curr_ch % 1000)
            pnt = kml_fit.newpoint(
                name=f"{km}+{m:03d}",
                coords=[(smoothed_coords[0][1], smoothed_coords[0][0])],
            )
            pnt.style.iconstyle.scale = 1.0
            pnt.style.iconstyle.color = simplekml.Color.red

            next_target = curr_ch + minor_interval

            for idx in range(len(smoothed_coords) - 1):
                p1 = smoothed_coords[idx]
                p2 = smoothed_coords[idx + 1]
                seg_m = geodesic(p1, p2).meters

                while (curr_ch + accumulated_dist + seg_m) >= next_target:
                    overshoot = next_target - (curr_ch + accumulated_dist)
                    frac = overshoot / seg_m if seg_m > 0 else 0

                    t_lat = p1[0] + frac * (p2[0] - p1[0])
                    t_lon = p1[1] + frac * (p2[1] - p1[1])

                    km = int(next_target // 1000)
                    m = int(next_target % 1000)
                    ch_text = f"{km}+{m:03d}"

                    ch_pnt = kml_fit.newpoint(
                        name=ch_text, coords=[(t_lon, t_lat)]
                    )
                    if int(next_target) % major_interval == 0:
                        ch_pnt.style.iconstyle.scale = 1.0
                        ch_pnt.style.iconstyle.color = simplekml.Color.red
                    else:
                        ch_pnt.style.iconstyle.scale = 0.6
                        ch_pnt.style.iconstyle.color = simplekml.Color.yellow

                    next_target += minor_interval

                accumulated_dist += seg_m

            total_road_length = accumulated_dist

            st.success(
                f"✅ Success! Generated Best Fit Alignment & Tour for {total_road_length/1000:.3f} km"
            )

            df_curves = pd.DataFrame(curve_data)
            st.write("### 📋 Horizontal Curve Geometry Schedule")
            st.dataframe(df_curves, use_container_width=True)

            col_dl1, col_dl2, col_dl3, col_dl4 = st.columns(4)

            fname_out = (
                output_name.strip()
                if output_name.strip()
                else "Alignment_With_Chainages"
            )
            fname_out = (
                fname_out if fname_out.endswith(".kml") else f"{fname_out}.kml"
            )

            with col_dl1:
                st.download_button(
                    label="📥 Download Standard KML",
                    data=kml_fit.kml(),
                    file_name=fname_out,
                    mime="application/vnd.google-earth.kml+xml",
                )

            with col_dl2:
                voice_kml_data = generate_voice_tour_kml(
                    smoothed_coords, start_chainage
                )
                st.download_button(
                    label="🔊 Download Voice Alert Tour KML",
                    data=voice_kml_data,
                    file_name=f"Voice_Tour_{fname_out}",
                    mime="application/vnd.google-earth.kml+xml",
                )

            with col_dl3:
                csv_data = df_curves.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="📊 Download CSV Data",
                    data=csv_data,
                    file_name="Best_Fit_Curve_Schedule.csv",
                    mime="text/csv",
                )

            with col_dl4:
                excel_report = generate_excel_alignment_report(
                    curve_data,
                    total_road_length,
                    design_speed,
                    "Existing Alignment Best Fit",
                )
                st.download_button(
                    label="📄 Download Excel Report",
                    data=excel_report,
                    file_name="Best_Fit_Alignment_Engineering_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
